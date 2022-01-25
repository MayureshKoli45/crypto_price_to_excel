import requests
import json
from openpyxl import workbook, load_workbook
from datetime import date

class CryptoPrices:
    # Constructor 
    def __init__(self):
        # For this values refer coingecko official website https://www.coingecko.com/
        # https://api.coingecko.com/api/v3/simple/price?ids=matic-network&vs_currencies=INR
        # I have broken down the api link to make a constructor
        self.apiurl = "https://api.coingecko.com/api/v3/simple/price?ids=" # we will set crypto id later
        self.vs_currencies = "&vs_currencies="  # we will set fiat currency later
        self.headers = {"accept": "application/json"}

    def getCoinPrice(self, crypto_id, fiat_currency):
        # This function will return a coin price
        # There are crypto ids which you need to search on coingecko some ids are mentioned in the below link
        # concatenating url to form a request
        fiat_currency = fiat_currency.upper() # We need uppercase
        coin_url = self.apiurl + crypto_id + self.vs_currencies + fiat_currency 
        r = requests.get(coin_url, headers=self.headers)
        data = r.json()[crypto_id][fiat_currency.lower()] # We need lower case because json file return lower case key
        return data

    def getMultiplePrices(self, crypto_id_list, fiat_currency):
        # This function will returns multiple coin price list
        data = []
        for ids in crypto_id_list:
            coin_url = self.apiurl + ids + self.vs_currencies + fiat_currency
            r = requests.get(coin_url, headers=self.headers)
            data.append(r.json()[ids][fiat_currency.lower()])
        return data

    def transferToExcel(self, prices_list):
        # This function will transfer your list data to excel
        wb = load_workbook("D:\My Important Documents\Crypto.xlsx") # Make a workbook and specify its location

        ws = wb.active

        # specify rows and columns
        i = 2
        while True:
            col = "B"
            if ws[col + str(i)].value == None:
                break

            else:
                ws[col + str(i)] = prices_list[i - 2]

            i += 1 

        update_date = date.today()   
        ws["G2"] = update_date    

        wb.save("D:\My Important Documents\Crypto.xlsx")
        print("Successfull Execution")
 
# Driver code -->
if __name__ == '__main__':
    while True:
        permission = input("Please check or update the script before running it...\nDo you want to run this program ?\nPress 'Y' for 'Yes' Press 'N' for 'No' :-\n").upper()
        
        try:
        
            if permission == "Y":
                crypto = CryptoPrices() # Making object of class CryptoPrices

                # For multiple coin prices make a list
                crypto_list = ["algorand","apenft","avalanche-2","binancecoin","bitcoin","cardano","chiliz","crypto-com-chain",
                "decentraland","ethereum","gala","hedera-hashgraph","nucypher","okb","polkadot","matic-network","ripple","shiba-inu",
                "solana","stellar","the-sandbox","vechain","terra-luna"]

                # New list containing prices only
                prices = crypto.getMultiplePrices(crypto_list, "INR") 

                # This function will transfer data to excel
                crypto.transferToExcel(prices)
                break

            elif permission == "N":
                break

            else:
                print("Please input the given commands\n") 

        except:
            print("Something Went Wrong!!!\n")      